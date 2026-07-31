"""
informe_ventas.py — Informe diario de ventas (12 meses + mes en curso) y stock por SKU.

QUÉ GENERA
Un Excel con una fila por SKU: datos maestros (U por caja, Formato, Color), el stock
consolidado de hoy y las ventas mensuales de los últimos 12 meses ENTEROS más el mes en
curso (hoy 31-07-2026 -> jul-2025 a jul-2026, 13 columnas).

FUENTES
  · Ventas   : SQL Server, dbo.ventas. BRUTO = sólo Facturas (sin NC ni ND),
               Segmento COMERCIAL, Cantidad > 0. En CAJAS.
  · Stock    : stock_actual.csv (el mismo que usa el plan). Consolidado Traverso +
               Montaner, 3 bodegas despachables. Ya viene en CAJAS (UMED=CJ).
  · Maestros : forecast/data/Parametros_Informe_Ventas.xlsx (U por caja, Formato, Color).
               Vive aparte de mrp_sku_params porque 81 de los 322 SKU del informe NO
               están en el MRP (importados, maquila, bundles).

REGLA DE EXCLUSIÓN (definida con Germán)
Los SKU sin U por caja en el Excel de parámetros se EXCLUYEN del informe. Pero NO en
silencio: se devuelven en `excluidos` y el correo los lista. Si mañana aparece un SKU
nuevo con ventas y nadie lo carga, sus ventas desaparecerían del informe sin que nadie
se entere — el mismo falso-cero que ya evitamos en faltantes y en el vigía.

También se excluye la categoría OTROS (reciclaje, recuperación de gastos: no son
productos).

USO
    python3 informe_ventas.py                 # genera el Excel en /app/data/
    python3 informe_ventas.py --salida X.xlsx
"""

# ── Ruta de los módulos compartidos ──────────────────────────────────────────
# Este paquete vive en una subcarpeta, pero usa módulos que están en la raíz de la
# app (db.py, stock.py). Se agrega esa raíz al path para poder importarlos.
# TRAVERSO_APP_DIR permite override si el montaje cambia.
import os as _os
import sys as _sys
_APP_DIR = _os.environ.get("TRAVERSO_APP_DIR", "/app")
if _APP_DIR not in _sys.path:
    _sys.path.insert(0, _APP_DIR)


import argparse
import logging
import os
from datetime import date

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger("informe_ventas")

# El Excel de maestros vive junto a este código, no en /app/data: es un insumo
# propio del informe, no un parámetro del MRP. Se sube por scp (está gitignoreado).
_AQUI = _os.path.dirname(_os.path.abspath(__file__))
PARAMS_XLSX = _os.environ.get(
    "INFORME_VENTAS_PARAMS",
    _os.path.join(_AQUI, "Parametros_Informe_Ventas.xlsx"))
SALIDA_DIR = _os.environ.get("INFORME_VENTAS_SALIDA", _os.path.join(_AQUI, "salida"))

CATEGORIAS_EXCLUIDAS = {"OTROS"}

NAVY = "1A2D4D"
CELESTE = "C0DCF0"


def _rango_meses(hoy: date) -> tuple[date, list[str]]:
    """12 meses ENTEROS + el mes en curso.

    Ejemplo con hoy = 31-07-2026: desde 2025-07-01, columnas 2025-07 .. 2026-07 (13).
    """
    y, m = hoy.year, hoy.month
    y0 = y - 1
    desde = date(y0, m, 1)
    meses, yy, mm = [], y0, m
    while (yy, mm) <= (y, m):
        meses.append(f"{yy:04d}-{mm:02d}")
        mm += 1
        if mm == 13:
            mm, yy = 1, yy + 1
    return desde, meses


def _leer_maestros() -> "pd.DataFrame":
    import pandas as pd
    if not os.path.exists(PARAMS_XLSX):
        raise RuntimeError(
            f"No existe el Excel de parámetros del informe: {PARAMS_XLSX}. "
            "Subirlo con scp a forecast/data/.")
    df = pd.read_excel(PARAMS_XLSX, header=2, dtype={"SKU": str})
    df = df[df["SKU"].notna()].copy()
    df["sku"] = df["SKU"].astype(str).str.strip()
    df["upc"] = pd.to_numeric(df["U por caja"], errors="coerce")
    df["formato"] = df["Formato"].fillna("").astype(str).str.strip()
    df["color"] = df["Color"].fillna("").astype(str).str.strip()
    logger.info("Maestros: %d SKU (%d con U por caja, %d con Color).",
                len(df), int(df["upc"].notna().sum()), int((df["color"] != "").sum()))
    return df[["sku", "upc", "formato", "color"]]


def _leer_ventas(desde: date, hasta_excl: date) -> "pd.DataFrame":
    import pandas as pd
    from sqlalchemy import text
    from db import get_engine
    q = text("""
        SELECT
            [Codigo Articulo]           AS sku,
            MAX([Nombre Articulo])      AS nombre,
            MAX([Categ. Comercial])     AS categoria,
            FORMAT([Fecha], 'yyyy-MM')  AS mes,
            SUM([Cantidad])             AS cantidad
        FROM dbo.ventas
        WHERE [Segmento] = 'COMERCIAL'
          AND [Tipo Doc] = 'Factura'
          AND [Cantidad] > 0
          AND [Codigo Articulo] IS NOT NULL AND [Codigo Articulo] <> ''
          AND CAST([Fecha] AS DATE) >= :d1
          AND CAST([Fecha] AS DATE) <  :d2
        GROUP BY [Codigo Articulo], FORMAT([Fecha], 'yyyy-MM')
    """)
    with get_engine().connect() as c:
        res = c.execute(q, {"d1": desde.isoformat(), "d2": hasta_excl.isoformat()})
        df = pd.DataFrame(res.fetchall(),
                          columns=["sku", "nombre", "categoria", "mes", "cantidad"])
    df["sku"] = df["sku"].astype(str).str.strip()
    logger.info("Ventas: %d filas, %d SKU, meses %s..%s",
                len(df), df["sku"].nunique(),
                df["mes"].min() if len(df) else "-", df["mes"].max() if len(df) else "-")
    return df


def _leer_stock() -> tuple["pd.Series", "pd.Series", str]:
    """(stock_cj por sku, descripcion por sku, fecha del snapshot)."""
    import pandas as pd
    from stock import load_stock_parquet
    df = load_stock_parquet()
    if df.empty:
        raise RuntimeError("stock_actual.csv está vacío: ¿corrió el refresh del plan?")
    df["sku"] = df["sku"].astype(str).str.strip()
    fecha = str(pd.to_datetime(df["fecha_descarga"]).max().date())
    logger.info("Stock: %d filas, %d SKU, snapshot %s", len(df), df["sku"].nunique(), fecha)
    return (df.groupby("sku")["stock_unidades"].sum(),
            df.groupby("sku")["descripcion"].first(), fecha)


def construir(hoy: date | None = None) -> dict:
    """Arma la tabla del informe. Devuelve dict con df, meses, excluidos y metadatos."""
    import pandas as pd
    hoy = hoy or date.today()
    desde, meses = _rango_meses(hoy)
    hasta_excl = date(hoy.year, hoy.month, hoy.day)  # el WHERE es < :d2, así que suma hoy
    hasta_excl = pd.Timestamp(hoy).date() + pd.Timedelta(days=1)
    hasta_excl = date(hasta_excl.year, hasta_excl.month, hasta_excl.day)

    maestros = _leer_maestros()
    ventas = _leer_ventas(desde, hasta_excl)
    stk, desc_s, fecha_stock = _leer_stock()

    piv = (ventas.pivot_table(index="sku", columns="mes", values="cantidad",
                              aggfunc="sum", fill_value=0)
           .reindex(columns=meses, fill_value=0).round(0).astype(int))
    nombre_v = ventas.groupby("sku")["nombre"].first()
    categ_v = ventas.groupby("sku")["categoria"].first()

    skus = sorted(set(piv.index) | set(stk.index))
    df = pd.DataFrame(index=skus)
    df.index.name = "sku"
    df = df.join(piv).join(stk.rename("stock_cj").round(0).astype(int))
    df[meses] = df[meses].fillna(0).astype(int)
    df["stock_cj"] = df["stock_cj"].fillna(0).astype(int)
    df["nombre"] = [nombre_v.get(s) if s in nombre_v.index else desc_s.get(s, "")
                    for s in df.index]
    df["categoria"] = [categ_v.get(s) if s in categ_v.index else "(sin venta)"
                       for s in df.index]

    m = maestros.set_index("sku")
    df["upc"] = [m.at[s, "upc"] if s in m.index else None for s in df.index]
    df["formato"] = [m.at[s, "formato"] if s in m.index else "" for s in df.index]
    df["color"] = [m.at[s, "color"] if s in m.index else "" for s in df.index]
    df["total"] = df[meses].sum(axis=1)

    # ── Exclusiones, con registro de lo excluido ─────────────────────────────
    df = df.reset_index()
    fuera_cat = df[df["categoria"].isin(CATEGORIAS_EXCLUIDAS)].copy()
    df = df[~df["categoria"].isin(CATEGORIAS_EXCLUIDAS)]
    sin_upc = df[df["upc"].isna()].copy()
    df = df[df["upc"].notna()].copy()
    df["upc"] = df["upc"].astype(int)

    df = df.sort_values(["categoria", "total"], ascending=[True, False]).reset_index(drop=True)

    excluidos = (sin_upc[["sku", "nombre", "categoria", "stock_cj", "total"]]
                 .sort_values("total", ascending=False).to_dict("records"))
    logger.info("Informe: %d SKU | excluidos por falta de U por caja: %d | "
                "excluidos por categoría: %d", len(df), len(sin_upc), len(fuera_cat))

    return {"df": df, "meses": meses, "excluidos": excluidos,
            "n_cat_excluidas": len(fuera_cat), "fecha_stock": fecha_stock,
            "desde": desde.isoformat(), "hoy": hoy.isoformat(),
            "total_ventas": int(df["total"].sum()),
            "total_stock": int(df["stock_cj"].sum())}


def generar_excel(rep: dict, salida: str | None = None) -> str:
    """Escribe el Excel y devuelve la ruta."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    df, meses = rep["df"], rep["meses"]
    os.makedirs(SALIDA_DIR, exist_ok=True)
    salida = salida or os.path.join(
        SALIDA_DIR, f"Informe_Ventas_Stock_{rep['hoy']}.xlsx")

    f_title = Font(name="Arial", size=14, bold=True, color=NAVY)
    f_sub = Font(name="Arial", size=9, italic=True, color="555555")
    f_hdr = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    f_data = Font(name="Arial", size=10)
    f_datab = Font(name="Arial", size=10, bold=True)
    fill_hdr = PatternFill("solid", fgColor=NAVY)
    fill_alt = PatternFill("solid", fgColor="F2F7FC")
    fill_tot = PatternFill("solid", fgColor=CELESTE)
    thin = Side(style="thin", color="D0D7E2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    NUM = "#,##0"

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas + Stock"
    cols = (["Cod SAP SKU", "Nombre SKU", "Categoría", "U por caja", "Formato", "Color",
             f"Stock {rep['fecha_stock']} (cj)"] + meses + ["Total período (cj)"])
    ncol = len(cols)
    C_M1 = 8
    C_M2 = C_M1 + len(meses) - 1
    C_TOT = ncol

    ws.cell(1, 1, "TRAVERSO S.A. — Ventas por mes y stock por SKU").font = f_title
    ws.cell(2, 1, (f"Ventas {meses[0]} a {meses[-1]} ({len(meses)-1} meses enteros + mes "
                   f"en curso) · BRUTO: sólo Facturas, sin NC ni ND · Segmento COMERCIAL "
                   f"· en CAJAS · Stock al {rep['fecha_stock']} (Traverso + Montaner, "
                   f"3 bodegas) · Excluye categoría OTROS y SKU sin U por caja "
                   f"· generado {rep['hoy']}")).font = f_sub

    HDR = 3
    for j, c in enumerate(cols, 1):
        cell = ws.cell(HDR, j, c)
        cell.font = f_hdr
        cell.fill = fill_hdr
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    r0 = HDR + 1
    for i, row in df.iterrows():
        r = r0 + i
        alt = (i % 2 == 1)
        vals = ([row["sku"], row["nombre"], row["categoria"], int(row["upc"]),
                 row["formato"], row["color"], int(row["stock_cj"])]
                + [int(row[m]) for m in meses])
        for j, v in enumerate(vals, 1):
            cell = ws.cell(r, j, v)
            cell.border = border
            cell.font = f_data
            if j >= 4:
                cell.number_format = NUM
                cell.alignment = Alignment(horizontal="right")
            if j in (1, 5, 6):
                cell.alignment = Alignment(horizontal="left")
            if alt:
                cell.fill = fill_alt
        t = ws.cell(r, C_TOT,
                    f"=SUM({get_column_letter(C_M1)}{r}:{get_column_letter(C_M2)}{r})")
        t.font = f_datab
        t.number_format = NUM
        t.border = border
        if alt:
            t.fill = fill_alt

    rt = r0 + len(df)
    ws.cell(rt, 1, "TOTAL").font = f_datab
    for j in range(1, ncol + 1):
        cell = ws.cell(rt, j)
        cell.fill = fill_tot
        cell.border = border
        if j == 7 or j >= C_M1:      # stock, meses y total (no U por caja/Formato/Color)
            L = get_column_letter(j)
            cell.value = f"=SUM({L}{r0}:{L}{rt-1})"
            cell.font = f_datab
            cell.number_format = NUM

    for col, w in zip("ABCDEFG", [13, 44, 17, 10, 9, 10, 16]):
        ws.column_dimensions[col].width = w
    for j in range(C_M1, C_M2 + 1):
        ws.column_dimensions[get_column_letter(j)].width = 9.5
    ws.column_dimensions[get_column_letter(C_TOT)].width = 14
    ws.freeze_panes = f"{get_column_letter(C_M1)}4"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[HDR].height = 30
    ws.auto_filter.ref = f"A{HDR}:{get_column_letter(ncol)}{rt-1}"

    # ── Hoja de excluidos: nunca en silencio ─────────────────────────────────
    if rep["excluidos"]:
        w2 = wb.create_sheet("Excluidos")
        w2.cell(1, 1, "SKU excluidos por falta de U por caja").font = f_title
        w2.cell(2, 1, ("Tienen venta o stock pero no están cargados en "
                       "Parametros_Informe_Ventas.xlsx. Sus ventas NO están en la hoja "
                       "principal. Cargarlos para que aparezcan.")).font = f_sub
        for j, c in enumerate(["Cod SAP SKU", "Nombre", "Categoría",
                               "Stock (cj)", "Ventas período (cj)"], 1):
            cell = w2.cell(3, j, c)
            cell.font = f_hdr
            cell.fill = fill_hdr
            cell.border = border
        for i, e in enumerate(rep["excluidos"]):
            for j, v in enumerate([e["sku"], e["nombre"], e["categoria"],
                                   int(e["stock_cj"]), int(e["total"])], 1):
                cell = w2.cell(4 + i, j, v)
                cell.font = f_data
                cell.border = border
                if j >= 4:
                    cell.number_format = NUM
        for col, w in zip("ABCDE", [13, 46, 18, 12, 18]):
            w2.column_dimensions[col].width = w
        w2.freeze_panes = "A4"

    wb.save(salida)
    logger.info("Excel generado: %s (%d SKU, %d excluidos)",
                salida, len(df), len(rep["excluidos"]))
    return salida


def ejecutar(hoy: date | None = None, salida: str | None = None) -> tuple[str, dict]:
    rep = construir(hoy=hoy)
    ruta = generar_excel(rep, salida=salida)
    return ruta, rep


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--salida", default=None, help="ruta del xlsx de salida")
    args = ap.parse_args()
    ruta, rep = ejecutar(salida=args.salida)
    print(f"\nOK -> {ruta}")
    print(f"SKU en el informe : {len(rep['df'])}")
    print(f"Excluidos sin UPC : {len(rep['excluidos'])}")
    print(f"Ventas del período: {rep['total_ventas']:,} cj")
    print(f"Stock ({rep['fecha_stock']}) : {rep['total_stock']:,} cj")
