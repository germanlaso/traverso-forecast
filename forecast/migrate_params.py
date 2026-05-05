"""
migrate_params.py
Migra los parámetros MRP del Excel a PostgreSQL.
Ejecutar una sola vez (o cuando se quiera reimportar desde el Excel).

Uso:
    python migrate_params.py [ruta_excel]
    python migrate_params.py /app/data/Traverso_Parametros_MRP.xlsx
"""
import sys
import math
import openpyxl
from db_mrp import (
    crear_tablas_params,
    upsert_linea,
    upsert_sku_params,
    upsert_sku_linea,
    borrar_todas_sku_lineas,
    borrar_todas_lineas,
    borrar_toda_setup_matrix,
    upsert_setup_entry,
    get_all_sku_lineas,
    get_session,
)
from sqlalchemy import text

EXCEL_PATH = sys.argv[1] if len(sys.argv) > 1 else "/app/data/Traverso_Parametros_MRP.xlsx"

def _float(v, default=0.0):
    try: return float(v) if v is not None else default
    except: return default

def _int(v, default=0):
    try: return int(float(v)) if v is not None else default
    except: return default

def _str(v, default=""):
    return str(v).strip() if v is not None else default

# V4: SKU_LINEA trae códigos en mayúsculas (SACHETERA, L1PET LV, L1PET A) pero
# mrp_lineas (cargada desde LINEAS_PRODUCCION y SKU_PARAMS) usa formato canónico.
# Normalizar al leer SKU_LINEA para que las claves matcheen entre tablas.
LINEA_CODIGO_NORMALIZER = {
    "SACHETERA": "Sachetera",
    "L1PET LV":  "L1Pet LV",
    "L1PET A":   "L1Pet A",
}


def migrar():
    print(f"Leyendo Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print(f"Pestañas: {wb.sheetnames}")

    # ── Crear tablas ──────────────────────────────────────────────────────────
    crear_tablas_params()
    print("Tablas creadas/verificadas en PostgreSQL")

    # ── Migrar LINEAS_PRODUCCION ──────────────────────────────────────────────
    ws_lin = wb['LINEAS_PRODUCCION']
    rows_lin = list(ws_lin.iter_rows(min_row=4, values_only=True))

    # Limpieza previa: borrar primero las hijas (mrp_sku_lineas) y después la
    # padre (mrp_lineas). No hay FK declarada hoy, pero el orden es buena
    # práctica defensiva por si se agrega en el futuro.
    borrar_todas_sku_lineas()
    borrar_todas_lineas()

    n_lineas = 0
    codigos_vistos = set()
    for row in rows_lin:
        codigo = _str(row[0])
        if not codigo or codigo.startswith("Nota"):
            continue
        # V4: LINEAS_PRODUCCION trae 1 fila por par (línea, SKU). Deduplicar
        # para insertar cada línea una sola vez en mrp_lineas.
        if codigo in codigos_vistos:
            continue
        codigos_vistos.add(codigo)
        activa = _str(row[10], "S").upper() == "S"
        if not activa:
            continue

        turnos    = _int(row[3], 1)
        horas     = _float(row[4], 8)
        dias      = _int(row[5], 5)
        velocidad = _float(row[7], 0)

        linea = {
            "codigo":         codigo,
            "nombre":         _str(row[1]),
            "area":           _str(row[2]),
            "turnos_dia":     turnos,
            "horas_turno":    horas,
            "dias_semana":    dias,
            "velocidad_u_hr": velocidad,
            "activa":         True,
        }
        upsert_linea(linea)
        cap = turnos * horas * dias * velocidad
        print(f"  Línea {codigo}: {_str(row[1])} | vel={velocidad} u/hr | cap={cap:,.0f} u/sem")
        n_lineas += 1

    print(f"→ {n_lineas} líneas migradas")

    # ── Migrar SKU_PARAMS ─────────────────────────────────────────────────────
    ws_sku = wb['SKU_PARAMS']
    rows_sku = list(ws_sku.iter_rows(min_row=4, values_only=True))

    n_skus = 0
    for row in rows_sku:
        sku = _str(row[0])
        if not sku or not sku[0].isdigit():
            continue

        activo_val = _str(row[14], "S").upper()
        activo = activo_val == "S"

        # V4: la columna 10 trae el código de línea directo (Sachetera, L1Pet LV,
        # L1Pet A). IMPORTACION viene vacía → linea_cod queda en "".
        linea_cod = _str(row[10])

        params = {
            "sku":             sku,
            "descripcion":     _str(row[1]),
            "categoria":       _str(row[3]),
            "tipo":            _str(row[4], "PRODUCCION"),
            "u_por_caja":      _int(row[2], 1),
            "lead_time_sem":   _float(row[5], 1),
            "ss_dias":         _int(row[6], 15),
            "batch_min_u":     _int(row[7], 0),
            "batch_mult_u":    _int(row[8], 1) or 1,
            "cap_bodega_u":    _int(row[9], 999999) or 999999,
            "t_cambio_hrs":    _float(row[11], 0),
            "linea_preferida": linea_cod,
            "activo":          activo,
        }
        upsert_sku_params(params)
        print(f"  SKU {sku}: {params['descripcion'][:35]} | linea={linea_cod} | cap_bod={params['cap_bodega_u']:,}")
        n_skus += 1

    print(f"→ {n_skus} SKUs migrados")

    # ── Migrar SKU_LINEA (pares SKU ↔ Línea con t_cambio y factor_velocidad) ──
    # Estrategia: borrar todos los pares y re-insertar desde el Excel (fuente
    # de verdad). Así, si en el Excel se eliminó un par, también se elimina en BD.
    if 'SKU_LINEA' in wb.sheetnames:
        ws_sl = wb['SKU_LINEA']
        rows_sl = list(ws_sl.iter_rows(min_row=4, values_only=True))

        n_sku_lineas = 0
        for row in rows_sl:
            sku    = _str(row[0])      # col A: SKU
            linea  = _str(row[2])      # col C: Código Línea (V4 trae mayúsculas)
            linea  = LINEA_CODIGO_NORMALIZER.get(linea, linea)
            t_camb = _float(row[4], 0) # col E: T. Cambio Batch (hrs)
            pref_s = _str(row[5], "N").upper()  # col F: Línea Preferida (S/N)
            factor = _float(row[7], 1.0)        # col H: Factor_Linea
            # NOTA V4: col[6] es 'Notas' (texto libre), Factor_Linea está en col[7].
            # En V3 estaba en col[6]. Si vuelve a moverse, ajustar acá.

            # Filtrar filas vacías o de notas
            if not sku or not sku[0].isdigit():
                continue
            if not linea:
                print(f"  ⚠ SKU_LINEA fila ignorada (linea vacía): sku={sku}")
                continue
            if factor <= 0:
                print(f"  ⚠ factor_velocidad inválido para {sku}/{linea}, usando 1.0")
                factor = 1.0

            preferida = (pref_s == "S")

            rec = {
                "sku":              sku,
                "linea":            linea,
                "t_cambio_hrs":     t_camb,
                "preferida":        preferida,
                "factor_velocidad": factor,
            }
            upsert_sku_linea(rec)
            flag = " ★" if preferida else "  "
            warn = "  ⚠" if factor != 1.0 else ""
            print(f"  {flag} {sku} → {linea} | t_camb={t_camb}h | factor={factor}{warn}")
            n_sku_lineas += 1

        print(f"→ {n_sku_lineas} pares SKU-Línea migrados")
    else:
        print("⚠ Pestaña SKU_LINEA no encontrada en el Excel — paso saltado")

    _print_resumen()

    print("\n✅ Migración completada exitosamente")


def _print_resumen():
    """Conteo de filas por tabla post-carga. Sanity check visual."""
    print("\n=== Resumen post-carga ===")
    with get_session() as session:
        for tabla in ['mrp_lineas', 'mrp_sku_params', 'mrp_sku_lineas']:
            count = session.execute(text(f"SELECT COUNT(*) FROM {tabla}")).scalar()
            print(f"  {tabla}: {count} filas")


def cargar_setup_matrix_inicial():
    """
    Genera la matriz inicial simétrica desde mrp_sku_lineas.
    Para cada línea, para cada par (sku_a, sku_b) de los SKUs asignados a esa línea:
      - sku_a → sku_a: 0 (auto-transición)
      - sku_a → sku_b: t_cambio_hrs[sku_b, linea]  (predecesor anónimo)
    Cuando llegue la matriz real del Gerente (F5), esto se reemplaza por UPDATE.
    """
    print("\n=== Cargando mrp_setup_matrix (simétrica inicial) ===")
    borrar_toda_setup_matrix()

    sku_lineas = get_all_sku_lineas()  # [{sku, linea, t_cambio_hrs, ...}]

    por_linea = {}
    for sl in sku_lineas:
        por_linea.setdefault(sl["linea"], []).append(sl)

    total_filas = 0
    for linea, items in por_linea.items():
        t_cambio_de = {it["sku"]: float(it["t_cambio_hrs"]) for it in items}
        skus = list(t_cambio_de.keys())
        for sku_a in skus:
            for sku_b in skus:
                tiempo = 0.0 if sku_a == sku_b else t_cambio_de[sku_b]
                upsert_setup_entry(sku_a, sku_b, linea, tiempo)
                total_filas += 1
        print(f"  {linea}: {len(skus)} SKUs → {len(skus)**2} filas")

    print(f"Total filas insertadas: {total_filas}")


if __name__ == "__main__":
    migrar()
    cargar_setup_matrix_inicial()
