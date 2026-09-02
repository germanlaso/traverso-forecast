"""
faltantes_excel.py — Genera el Excel del "Informe de Quiebres de Stock Facturado"
para un día, replicando el formato del informe manual de Traverso (Planificación).

Dos pestañas:
  - "Informe": encabezado + KPIs del día + tabla RESUMEN POR SKU (Producto, Cod SAP,
    Causa, Stock, Programado, Faltante, %). % = faltante del SKU / total del día.
  - "Detalle por cliente": Producto, Cod SAP, Cliente, Causa, Faltante.

Lee los datos ya persistidos vía db_mrp.get_faltantes_por_fecha(fecha) (solo lectura).
Devuelve los bytes del .xlsx (para servir por el endpoint) o lo escribe a un path.
"""

import io
from datetime import date, datetime
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Paleta corporativa (del informe de referencia)
AZUL   = "1A2D4D"
CELESTE = "C0DCF0"
GRIS   = "F2F4F7"
GRIS_TXT = "4A5568"
BLANCO = "FFFFFF"

CAUSA_LABEL = {"sin_stock": "SIN STOCK", "vu_insuficiente": "VU INSUFICIENTE"}

# Agrupación del informe (debe coincidir con faltantes.py)
GRUPO_PRODUCCION = "Producción"
GRUPO_IMPORTACION = "Importación"
ORDEN_GRUPO = {GRUPO_PRODUCCION: 0, GRUPO_IMPORTACION: 1}


def _grupo_de(row):
    """Grupo de una fila; default Producción si falta o es desconocido."""
    g = (row.get("grupo") or "").strip()
    return g if g in ORDEN_GRUPO else GRUPO_PRODUCCION


def _miles(n):
    """Formato de miles con punto (es-CL): 2519 -> '2.519'."""
    return f"{int(round(n)):,}".replace(",", ".")

_thin = Side(style="thin", color="D3D1C7")
BORDE = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _f(size=9, bold=False, color=AZUL, name="Arial"):
    return Font(name=name, size=size, bold=bold, color=color)


def _fill(hexcolor):
    return PatternFill("solid", fgColor=hexcolor)


def _fecha_txt(f):
    if isinstance(f, (date, datetime)):
        return f.strftime("%d-%m-%Y")
    # 'YYYY-MM-DD' -> 'DD-MM-YYYY'
    p = str(f).split("-")
    return f"{p[2]}-{p[1]}-{p[0]}" if len(p) == 3 else str(f)


def _repo_txt(rp):
    """Texto de la fecha de reposición según el tipo (Faltantes V2).
    rp: {tipo, valor}. tipo ∈ auto|inactivo|sin_of|manual."""
    if not rp:
        return ""
    tipo = rp.get("tipo")
    valor = rp.get("valor")
    if tipo == "auto":
        return _fecha_txt(valor) if valor else ""
    if tipo == "inactivo":
        return "SKU inactivo"
    if tipo == "sin_of":
        return "Sin OF futura"
    if tipo == "manual":
        return _fecha_txt(valor) if valor else "—"
    return ""


def construir(fecha, filas, explicaciones=None, soluciones=None, reposicion=None, con_v2=False, data30=None):
    """fecha: str YYYY-MM-DD | date. filas: list de dicts de get_faltantes_por_fecha.
    explicaciones: dict opcional {sku: {explicacion, autor}} — si se pasa, agrega la
    columna 'Explicación' a la tabla resumen por SKU.
    soluciones: dict opcional {sku: {solucion, solucion_autor}} — Faltantes V2.
    reposicion: dict opcional {sku: {tipo, valor}} — Faltantes V2 (fecha de reposición).
    con_v2: si True, agrega las columnas 'Reposición' y 'Solución' (Faltantes V2).
    Devuelve un openpyxl Workbook."""
    fecha_str = fecha if isinstance(fecha, str) else fecha.isoformat()
    explicaciones = explicaciones or {}
    soluciones = soluciones or {}
    reposicion = reposicion or {}

    # --- agregación por SKU ---
    porsku = {}
    for r in filas:
        sku = r["sku"]
        g = porsku.get(sku)
        if g is None:
            g = {"sku": sku, "descripcion": r.get("descripcion", ""),
                 "stock_ini_cj": r.get("stock_ini_cj", 0),
                 "programado_cj": r.get("programado_cj", 0),
                 "faltante_cj": 0.0, "causas": set(), "n_cli": 0,
                 "grupo": _grupo_de(r)}
            porsku[sku] = g
        g["faltante_cj"] += float(r.get("faltante_cj", 0) or 0)
        g["n_cli"] += 1
        if r.get("causa"):
            g["causas"].add(r["causa"])
    # orden: grupo (Producción primero), luego faltante desc
    resumen = sorted(porsku.values(),
                     key=lambda x: (ORDEN_GRUPO.get(x["grupo"], 0), -x["faltante_cj"]))

    total_cj   = sum(g["faltante_cj"] for g in resumen)
    n_prod     = len(resumen)
    n_lineas   = len(filas)
    n_causas   = len({r.get("causa") for r in filas if r.get("causa")})

    wb = openpyxl.Workbook()

    # ══════════════════════ Pestaña "Informe" ══════════════════════
    ws = wb.active
    ws.title = "Informe"
    ws.sheet_view.showGridLines = False
    for col, w in zip("BCDEFGHI", [2, 40, 14, 22, 13, 14, 13, 9]):
        ws.column_dimensions[col].width = w

    # Encabezado
    ws.merge_cells("C3:I3")
    ws["C3"] = "TRAVERSO · Planificación"
    ws["C3"].font = _f(9, True, BLANCO); ws["C3"].fill = _fill(AZUL)
    ws["C3"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("C4:I4")
    ws["C4"] = "Informe de Quiebres de Stock Facturado"
    ws["C4"].font = _f(16, True, BLANCO); ws["C4"].fill = _fill(AZUL)
    ws["C4"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 26

    ws["C6"] = "Día del informe:"; ws["C6"].font = _f(9, True)
    ws["D6"] = _fecha_txt(fecha_str); ws["D6"].font = _f(9)
    ws["F6"] = "Fecha de emisión:"; ws["F6"].font = _f(9, True)
    ws["G6"] = _fecha_txt(date.today()); ws["G6"].font = _f(9)

    # Resumen del día
    ws["C8"] = "Resumen del día"; ws["C8"].font = _f(12, True)

    def kpi(rango_val, rango_lbl, valor, etiqueta):
        # rango_val/rango_lbl: p.ej. "C9:D9" / "C10:D10" (merge + centrar)
        ws.merge_cells(rango_val); ws.merge_cells(rango_lbl)
        cv = rango_val.split(":")[0]; cl = rango_lbl.split(":")[0]
        ws[cv] = valor
        ws[cv].font = _f(22, True); ws[cv].fill = _fill(CELESTE)
        ws[cv].alignment = Alignment(horizontal="center", vertical="center")
        ws[cl] = etiqueta
        ws[cl].font = _f(8, True, GRIS_TXT); ws[cl].fill = _fill(CELESTE)
        ws[cl].alignment = Alignment(horizontal="center", vertical="center")
        # pintar el resto de las celdas del merge (para que el fill cubra todo el ancho)
        for rng in (rango_val, rango_lbl):
            ini, fin = rng.split(":")
            col_i = openpyxl.utils.column_index_from_string(ini[0])
            col_f = openpyxl.utils.column_index_from_string(fin[0])
            fila = int(ini[1:])
            for cc in range(col_i, col_f + 1):
                ws.cell(fila, cc).fill = _fill(CELESTE)

    # C:D | E:F | G:I (este último 3 celdas, para que "CAJAS CON QUIEBRE" no desborde)
    kpi("C9:D9", "C10:D10", n_prod, "PRODUCTOS CON QUIEBRE")
    kpi("E9:F9", "E10:F10", n_lineas, "LÍNEAS (SKU × CLIENTE)")
    kpi("G9:I9", "G10:I10", int(round(total_cj)), "CAJAS CON QUIEBRE")
    ws.row_dimensions[9].height = 34

    # Tabla resumen por SKU
    hdr_row = 12
    con_expl = bool(explicaciones)
    headers = ["Producto", "Cod. SAP", "Causa", "Stock (cj)", "Programado (cj)", "Faltante (cj)", "%"]
    # Índices de columnas variables (para alineación por columna, no por posición fija)
    idx_repo = idx_expl = idx_sol = None
    if con_v2:
        # V2: Reposición | Explicación | Solución (las 3 de gestión juntas al final)
        idx_repo = len(headers); headers.append("Reposición")
        idx_expl = len(headers); headers.append("Explicación")
        idx_sol  = len(headers); headers.append("Solución")
        ws.column_dimensions["J"].width = 16   # Reposición
        ws.column_dimensions["K"].width = 40   # Explicación
        ws.column_dimensions["L"].width = 40   # Solución
    elif con_expl:
        idx_expl = len(headers); headers.append("Explicación")
        ws.column_dimensions["J"].width = 45   # columna de explicación
    for i, h in enumerate(headers):
        cell = ws.cell(hdr_row, 3 + i, h)
        cell.font = _f(9, True); cell.fill = _fill(CELESTE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDE

    # subtotales por grupo (para las bandas de sección)
    subtot = defaultdict(float)
    for g in resumen:
        subtot[g["grupo"]] += g["faltante_cj"]

    def _banda(row, texto):
        c1, c2 = 3, 3 + len(headers) - 1
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        for cc in range(c1, c2 + 1):
            b = ws.cell(row, cc); b.fill = _fill(AZUL); b.border = BORDE
        cell = ws.cell(row, c1, texto)
        cell.font = _f(9, True, BLANCO)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18

    r = hdr_row + 1
    grupo_actual = None
    zebra = 0
    for g in resumen:
        if g["grupo"] != grupo_actual:
            grupo_actual = g["grupo"]
            _banda(r, f'{grupo_actual.upper()}  —  {_miles(subtot[grupo_actual])} cj')
            r += 1
            zebra = 0
        causas = g["causas"]
        causa_txt = "MIXTA" if len(causas) > 1 else (CAUSA_LABEL.get(next(iter(causas)), "") if causas else "")
        pct = (g["faltante_cj"] / total_cj) if total_cj else 0
        vals = [g["descripcion"], g["sku"], causa_txt,
                round(g["stock_ini_cj"], 0), round(g["programado_cj"], 0),
                round(g["faltante_cj"], 0), pct]
        if con_v2:
            # Reposición (texto según tipo)
            vals.append(_repo_txt(reposicion.get(g["sku"])))
            # Explicación
            ex = explicaciones.get(g["sku"], {})
            txt_e = (ex.get("explicacion") or "").strip()
            aut_e = (ex.get("autor") or "").strip()
            vals.append(f"{txt_e}  ({aut_e})" if (txt_e and aut_e) else txt_e)
            # Solución
            so = soluciones.get(g["sku"], {})
            txt_s = (so.get("solucion") or "").strip()
            aut_s = (so.get("solucion_autor") or "").strip()
            vals.append(f"{txt_s}  ({aut_s})" if (txt_s and aut_s) else txt_s)
        elif con_expl:
            ex = explicaciones.get(g["sku"], {})
            txt = (ex.get("explicacion") or "").strip()
            autor = (ex.get("autor") or "").strip()
            vals.append(f"{txt}  ({autor})" if (txt and autor) else txt)
        for i, v in enumerate(vals):
            cell = ws.cell(r, 3 + i, v)
            cell.font = _f(8, False, "1A2332", name="Calibri")
            cell.fill = _fill(GRIS if zebra % 2 else BLANCO)
            cell.border = BORDE
            if i == 0:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            elif i in (3, 4, 5):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0"
            elif i == 6:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = "0.0%"
            elif i == idx_repo:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif i in (idx_expl, idx_sol):
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        r += 1
        zebra += 1

    # fila total general
    ws.cell(r, 3, "TOTAL").font = _f(9, True)
    tc = ws.cell(r, 8, round(total_cj, 0)); tc.font = _f(9, True)
    tc.number_format = "#,##0"; tc.alignment = Alignment(horizontal="right")

    # ══════════════════════ Pestaña "Detalle por cliente" ══════════════════════
    ws2 = wb.create_sheet("Detalle por cliente")
    ws2.sheet_view.showGridLines = False
    for col, w in zip("BCDEFG", [2, 40, 14, 34, 18, 14]):
        ws2.column_dimensions[col].width = w

    ws2.merge_cells("C2:G2")
    ws2["C2"] = f"Detalle por cliente · {_fecha_txt(fecha_str)}"
    ws2["C2"].font = _f(13, True, BLANCO); ws2["C2"].fill = _fill(AZUL)
    ws2["C2"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[2].height = 22

    # KPIs de la pestaña detalle (mismos valores del día)
    def kpi2(rango_val, rango_lbl, valor, etiqueta):
        ws2.merge_cells(rango_val); ws2.merge_cells(rango_lbl)
        cv = rango_val.split(":")[0]; cl = rango_lbl.split(":")[0]
        ws2[cv] = valor
        ws2[cv].font = _f(20, True); ws2[cv].alignment = Alignment(horizontal="center", vertical="center")
        ws2[cl] = etiqueta
        ws2[cl].font = _f(8, True, GRIS_TXT); ws2[cl].alignment = Alignment(horizontal="center", vertical="center")
        for rng in (rango_val, rango_lbl):
            ini, fin = rng.split(":")
            ci = openpyxl.utils.column_index_from_string(ini[0]); cf = openpyxl.utils.column_index_from_string(fin[0])
            fila = int(ini[1:])
            for cc in range(ci, cf + 1):
                ws2.cell(fila, cc).fill = _fill(CELESTE)
    kpi2("C4:C4", "C5:C5", n_prod, "PRODUCTOS")
    kpi2("D4:E4", "D5:E5", n_lineas, "LÍNEAS (SKU × CLIENTE)")
    kpi2("F4:G4", "F5:G5", int(round(total_cj)), "CAJAS CON QUIEBRE")
    ws2.row_dimensions[4].height = 30

    h2 = ["Producto", "Cod. SAP", "Cliente", "Causa", "Faltante (cj)"]
    for i, h in enumerate(h2):
        cell = ws2.cell(7, 3 + i, h)
        cell.font = _f(9, True); cell.fill = _fill(CELESTE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDE

    # orden: grupo (Producción primero), luego faltante desc
    filas_ord = sorted(filas, key=lambda x: (ORDEN_GRUPO.get(_grupo_de(x), 0),
                                             -float(x.get("faltante_cj", 0) or 0), x["sku"]))
    subtot2 = defaultdict(float)
    for x in filas:
        subtot2[_grupo_de(x)] += float(x.get("faltante_cj", 0) or 0)

    def _banda2(row, texto):
        c1, c2 = 3, 3 + len(h2) - 1
        ws2.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        for cc in range(c1, c2 + 1):
            b = ws2.cell(row, cc); b.fill = _fill(AZUL); b.border = BORDE
        cell = ws2.cell(row, c1, texto)
        cell.font = _f(9, True, BLANCO)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[row].height = 18

    rr = 8
    grupo_actual = None
    zebra = 0
    for r0 in filas_ord:
        gr = _grupo_de(r0)
        if gr != grupo_actual:
            grupo_actual = gr
            _banda2(rr, f'{gr.upper()}  —  {_miles(subtot2[gr])} cj')
            rr += 1
            zebra = 0
        causa_txt = CAUSA_LABEL.get(r0.get("causa"), r0.get("causa", ""))
        vals = [r0.get("descripcion", ""), r0["sku"], r0.get("nom_cliente", ""),
                causa_txt, round(float(r0.get("faltante_cj", 0) or 0), 0)]
        for i, v in enumerate(vals):
            cell = ws2.cell(rr, 3 + i, v)
            cell.font = _f(8, False, "1A2332", name="Calibri")
            cell.fill = _fill(GRIS if zebra % 2 else BLANCO)
            cell.border = BORDE
            if i == 4:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0"
            elif i in (0, 2, 3):
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        rr += 1
        zebra += 1

    # Hoja "Resumen 30d" (opcional): si viene data30, se agrega como pestaña aparte.
    # Degrada con gracia: si algo falla, el Excel se genera igual sin esa hoja.
    if data30 and (data30.get("filas")):
        try:
            hoja_resumen30(wb, data30)
        except Exception:
            pass

    return wb


def generar_bytes(fecha, filas, explicaciones=None, soluciones=None, reposicion=None, con_v2=False, data30=None):
    """Devuelve los bytes del .xlsx (para servir por HTTP)."""
    wb = construir(fecha, filas, explicaciones, soluciones, reposicion, con_v2, data30=data30)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generar_bytes_resumen30(data30):
    """Devuelve los bytes de un .xlsx con SOLO la hoja Resumen 30d (para descarga)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hoja_resumen30(wb, data30)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


if __name__ == "__main__":
    # Prueba: genera el informe de un día a un archivo local.
    import sys
    from db_mrp import get_faltantes_por_fecha
    f = sys.argv[1] if len(sys.argv) > 1 else str(date.today())
    filas = get_faltantes_por_fecha(f)
    wb = construir(f, filas)
    out = f"/tmp/informe_faltantes_{f}.xlsx"
    wb.save(out)
    print(f"OK: {out} | {len(filas)} filas | {len(set(r['sku'] for r in filas))} SKU")


# ══════════════════════ Hoja "Resumen 30d" ══════════════════════
# Recibe el dict de db_mrp.resumen_faltantes_30d(): {desde, hasta, fechas[],
# no_habil[], filas[{sku,descripcion,grupo,cat_comercial,ventas_30,faltante_30,
# pct,dias_con_falta,serie[]}], resumen_cat{cat:{grupo,ventas_total,n_sin_falta}}}.

_HDR_NOHABIL = "EAE6DA"   # fondo de encabezado para sábado/domingo/feriado

def _color_celda_xlsx(cj):
    """Relleno de celda según cajas faltantes del día (misma escala que el dashboard)."""
    if not cj or cj <= 0:
        return None
    if cj <= 20:  return "F7D65A"   # amarillo
    if cj <= 60:  return "EF9F27"   # naranja
    return "E24B4A"                 # rojo


def _fecha_ddmm(iso):
    p = str(iso).split("-")
    return f"{p[2]}-{p[1]}" if len(p) == 3 else str(iso)


def hoja_resumen30(wb, data30, titulo="Resumen 30d"):
    """Agrega al workbook `wb` una hoja con el resumen de faltantes de 30 días.
    Devuelve la hoja creada."""
    fechas   = data30.get("fechas", []) or []
    no_hab   = data30.get("no_habil", []) or []
    filas    = data30.get("filas", []) or []
    rcat     = data30.get("resumen_cat", {}) or {}
    nd = len(fechas)

    ws = wb.create_sheet(titulo)
    ws.sheet_view.showGridLines = False
    # anchos: B margen, C Cod, D Producto, E Ventas, F Faltante, G %, H Días
    for col, w in zip("BCDEFGH", [2, 12, 34, 12, 12, 10, 7]):
        ws.column_dimensions[col].width = w
    # columnas de días (I en adelante): angostas
    for i in range(nd):
        ws.column_dimensions[get_column_letter(9 + i)].width = 4.2

    # ── Título + rango ──
    ult = 8 + nd
    ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=ult)
    t = ws.cell(2, 3, f"Resumen de faltantes — últimos 30 días  ({_fecha_ddmm(data30.get('desde',''))} al {_fecha_ddmm(data30.get('hasta',''))})")
    t.font = _f(13, True, BLANCO); t.fill = _fill(AZUL)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22

    # ── Encabezados de columna (fila 4) ──
    hdr_row = 4
    fijos = ["Cod", "Producto", "Ventas 30d", "Faltante 30d", "%", "Días"]
    for i, h in enumerate(fijos):
        c = ws.cell(hdr_row, 3 + i, h)
        c.font = _f(9, True); c.fill = _fill(CELESTE)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = BORDE
    for j, iso in enumerate(fechas):
        c = ws.cell(hdr_row, 9 + j, _fecha_ddmm(iso))
        nh = j < len(no_hab) and no_hab[j]
        c.font = _f(8, bool(nh), GRIS_TXT if nh else "4A5568")
        c.fill = _fill(_HDR_NOHABIL if nh else CELESTE)
        c.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90); c.border = BORDE
    ws.row_dimensions[hdr_row].height = 44

    # ── helpers de banda/fila ──
    def _num(cell, val, fmt="#,##0", bold=False, color="1A2332"):
        cell.value = val; cell.font = _f(8, bold, color, name="Calibri")
        cell.number_format = fmt; cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = BORDE

    def banda(row, label, ventas, falta, fill_hex, txt_hex, bold=True, sz=9):
        # etiqueta (C:D) + Ventas + Faltante + % + Días(vacío) + días(vacío)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        c = ws.cell(row, 3, label); c.font = _f(sz, bold, txt_hex); c.fill = _fill(fill_hex)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row, 4).fill = _fill(fill_hex)
        cv = ws.cell(row, 5, round(ventas)); cv.font = _f(sz, bold, txt_hex); cv.fill = _fill(fill_hex)
        cv.number_format = "#,##0"; cv.alignment = Alignment(horizontal="right", vertical="center")
        cf = ws.cell(row, 6, round(falta)); cf.font = _f(sz, bold, txt_hex); cf.fill = _fill(fill_hex)
        cf.number_format = "#,##0"; cf.alignment = Alignment(horizontal="right", vertical="center")
        pct = (falta / ventas * 100) if ventas else None
        cp = ws.cell(row, 7, (pct / 100.0) if pct is not None else "—"); cp.font = _f(sz, bold, txt_hex); cp.fill = _fill(fill_hex)
        if pct is not None: cp.number_format = "0.0%"
        cp.alignment = Alignment(horizontal="right", vertical="center")
        for cc in range(8, 9 + nd):
            ws.cell(row, cc).fill = _fill(fill_hex)

    # ── subtotales por categoría + ventas de "sin faltante" ──
    falta_cat = {}
    vent_falt_cat = {}
    for f in filas:
        falta_cat[f["cat_comercial"]] = falta_cat.get(f["cat_comercial"], 0) + (f["faltante_30"] or 0)
        vent_falt_cat[f["cat_comercial"]] = vent_falt_cat.get(f["cat_comercial"], 0) + (f["ventas_30"] or 0)

    def _grp_de_cat(cat):
        for f in filas:
            if f["cat_comercial"] == cat: return f["grupo"]
        return GRUPO_PRODUCCION
    subtot_grupo = {}
    for cat, rc in rcat.items():
        g = rc.get("grupo") or _grp_de_cat(cat)
        sg = subtot_grupo.setdefault(g, {"ventas": 0, "falta": 0})
        sg["ventas"] += (rc.get("ventas_total") or 0)
        sg["falta"]  += falta_cat.get(cat, 0)

    # ── recorrer filas con bandas ──
    r = hdr_row + 1
    grupo_actual = catActual = None
    zebra = 0

    def cerrar_categoria(cat):
        rc = rcat.get(cat, {})
        n_sf = rc.get("n_sin_falta", 0)
        vent_sf = (rc.get("ventas_total", 0) or 0) - vent_falt_cat.get(cat, 0)
        if n_sf > 0 and vent_sf > 0.5:
            nonlocal r
            cc = ws.cell(r, 3, f"({n_sf} SKU sin faltante)")
            cc.font = _f(8, False, "888780", name="Calibri"); cc.font = Font(name="Calibri", size=8, italic=True, color="888780")
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
            cc.alignment = Alignment(horizontal="left", vertical="center")
            _num(ws.cell(r, 5), round(vent_sf), color="888780")
            _num(ws.cell(r, 6), 0, color="888780")
            cp = ws.cell(r, 7, 0.0); cp.number_format = "0.0%"; cp.font = _f(8, False, "888780", name="Calibri")
            cp.alignment = Alignment(horizontal="right", vertical="center")
            r += 1

    for f in filas:
        if f["cat_comercial"] != catActual and catActual is not None:
            cerrar_categoria(catActual)
        if f["grupo"] != grupo_actual:
            grupo_actual = f["grupo"]; catActual = None; zebra = 0
            sg = subtot_grupo.get(f["grupo"], {"ventas": 0, "falta": 0})
            banda(r, f["grupo"].upper(), sg["ventas"], sg["falta"], AZUL, BLANCO, sz=10); r += 1
        if f["cat_comercial"] != catActual:
            catActual = f["cat_comercial"]; zebra = 0
            rc = rcat.get(f["cat_comercial"], {})
            vt = rc.get("ventas_total", 0) or 0
            banda(r, f["cat_comercial"] or "(sin categoría)", vt, falta_cat.get(f["cat_comercial"], 0),
                  GRIS, GRIS_TXT, sz=9); r += 1
        # fila SKU
        bg = GRIS if zebra % 2 else BLANCO; zebra += 1
        ws.cell(r, 3, f["sku"]).font = _f(8, True, "1D9E75", name="Calibri")
        cd = ws.cell(r, 4, f["descripcion"]); cd.font = _f(8, False, "1A2332", name="Calibri")
        cd.alignment = Alignment(horizontal="left", vertical="center")
        _num(ws.cell(r, 5), round(f["ventas_30"]) if f["ventas_30"] is not None else "—", color="888780")
        _num(ws.cell(r, 6), round(f["faltante_30"]), bold=True, color="E24B4A")
        cp = ws.cell(r, 7, (f["pct"] / 100.0) if f["pct"] is not None else "—")
        if f["pct"] is not None: cp.number_format = "0.0%"
        cp.font = _f(8, True, "1A2332", name="Calibri"); cp.alignment = Alignment(horizontal="right", vertical="center")
        cdd = ws.cell(r, 8, f["dias_con_falta"]); cdd.font = _f(8, False, "1A2332", name="Calibri")
        cdd.alignment = Alignment(horizontal="center", vertical="center")
        # celdas de día: número + color
        for j, cj in enumerate(f["serie"]):
            cell = ws.cell(r, 9 + j)
            hexcol = _color_celda_xlsx(cj)
            if hexcol:
                cell.value = round(cj); cell.fill = _fill(hexcol)
                cell.font = _f(7, False, "FFFFFF" if cj > 60 else "1A2332", name="Calibri")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for cc in range(3, 9):
            ws.cell(r, cc).fill = _fill(bg) if ws.cell(r, cc).fill.fgColor.rgb in (None, "00000000") else ws.cell(r, cc).fill
        r += 1
    if catActual is not None:
        cerrar_categoria(catActual)

    ws.freeze_panes = "I5"   # fija columnas fijas + encabezado
    return ws
