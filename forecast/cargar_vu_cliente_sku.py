"""
cargar_vu_cliente_sku.py — Carga la tabla de VU mínima por cliente × SKU de logística
(Excel, pestaña "Base datos") a la tabla Postgres mrp_vu_cliente_sku.

El dato que usa el motor de faltantes es `min_dias` (VU mínima absoluta en días para
despachar ese SKU a ese cliente). Se guarda también `min_meses` (referencia) y un
`pct_derivado = min_dias / (VU_total_del_maestro)` como referencia legible (NULL si el
SKU no tiene VU en el maestro de artículos).

Columnas esperadas en la pestaña "Base datos":
  Codigo SN | Nombre SN | SAP | DUN | DESCRIPCION | BASE | LAYER |
  Cajas por pallet | Minimo meses para despacho | Minimo dias para despacho

Uso:
    docker cp forecast/cargar_vu_cliente_sku.py traverso_forecast:/app/
    docker cp <excel> traverso_forecast:/app/Requerimiento_VU_clientes.xlsx
    docker exec traverso_forecast python3 /app/cargar_vu_cliente_sku.py /app/Requerimiento_VU_clientes.xlsx
"""

import sys
import logging
from datetime import date

import openpyxl
from sqlalchemy import text

from db import get_engine
from db_mrp import upsert_vu_cliente_sku

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

DIAS_POR_MES = 30
PESTANA = "Base datos"

COL = {"cli": "Codigo SN", "sku": "SAP", "desc": "DESCRIPCION",
       "meses": "Minimo meses para despacho", "dias": "Minimo dias para despacho"}


def _leer_vu_maestro(engine):
    """{sku: meses_vu} de PT='Y' para derivar el %."""
    vu = {}
    with engine.connect() as c:
        for sku, meses in c.execute(text(
                "SELECT LTRIM(RTRIM([ItemCode])), [MesDuracion] "
                "FROM dbo.MaestraArticuloV2 WHERE [PT]='Y'")).fetchall():
            s = str(sku).strip()
            m = int(meses) if meses is not None else 0
            if s not in vu or m > vu[s]:
                vu[s] = m
    return vu


def cargar(path_excel):
    wb = openpyxl.load_workbook(path_excel, read_only=True, data_only=True)
    if PESTANA not in wb.sheetnames:
        raise RuntimeError(f"No existe la pestaña '{PESTANA}'. Hay: {wb.sheetnames}")
    ws = wb[PESTANA]

    # mapear encabezados -> índice de columna
    hdr = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is not None:
            hdr[str(v).strip()] = c
    faltan = [v for v in COL.values() if v not in hdr]
    if faltan:
        raise RuntimeError(f"Faltan columnas en el Excel: {faltan}")

    vu_maestro = _leer_vu_maestro(get_engine())

    filas = []
    sin_vu = 0
    for r in range(2, ws.max_row + 1):
        cli = ws.cell(r, hdr[COL["cli"]]).value
        sku = ws.cell(r, hdr[COL["sku"]]).value
        dias = ws.cell(r, hdr[COL["dias"]]).value
        if cli is None or sku is None or dias is None:
            continue
        cod = str(cli).strip()
        s = str(sku).strip()
        min_dias = int(round(float(dias)))
        meses_v = ws.cell(r, hdr[COL["meses"]]).value
        min_meses = int(round(float(meses_v))) if meses_v is not None else None
        desc = ws.cell(r, hdr[COL["desc"]]).value
        desc = str(desc).strip() if desc else ""

        vu_total = vu_maestro.get(s, 0) * DIAS_POR_MES
        if vu_total > 0:
            pct = round(min_dias / vu_total, 4)
        else:
            pct = None
            sin_vu += 1

        filas.append({"cod_cliente": cod, "sku": s, "descripcion": desc[:120],
                      "min_dias": min_dias, "min_meses": min_meses, "pct_derivado": pct})

    n = upsert_vu_cliente_sku(filas)
    clientes = sorted({f["cod_cliente"] for f in filas})
    logger.info("Cargadas %d filas | %d clientes | %d SKU | %d sin VU en maestro (pct NULL).",
                n, len(clientes), len({f["sku"] for f in filas}), sin_vu)
    logger.info("Clientes: %s", ", ".join(clientes))
    # muestra
    print("Muestra (cliente, sku, min_dias, min_meses, pct_derivado):")
    for f in filas[:8]:
        print(f"  {f['cod_cliente']:<14} {f['sku']:<12} {f['min_dias']:>5}d "
              f"{str(f['min_meses']):>3}m  pct={f['pct_derivado']}")
    return n


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "/app/Requerimiento_VU_clientes.xlsx"
    total = cargar(path)
    print(f"\nOK: {total} filas en mrp_vu_cliente_sku.")
